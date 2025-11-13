from __future__ import annotations

from typing import List, Dict

from openpyxl import load_workbook

from ..models import ResultRow, Status


def read_contract_numbers(path: str) -> List[str]:
    result: List[str] = []
    seen = set()
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            s = line.strip()
            if not s or s.startswith("#"):
                continue
            if s in seen:
                continue
            seen.add(s)
            result.append(s)
    return result


def read_results_excel(path: str):
    wb = load_workbook(filename=path)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(x).strip() if x is not None else "" for x in header_row]
    idx = {name: i for i, name in enumerate(headers)}

    def norm(x):
        if x is None:
            return None
        s = str(x).strip()
        return s if s != "" else None

    order: List[str] = []
    mapping: Dict[str, ResultRow] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        cn = norm(row[idx.get("contract_number", -1)]) if idx.get("contract_number") is not None else None
        if not cn:
            continue
        cid = norm(row[idx.get("contract_id", -1)]) if idx.get("contract_id") is not None else None
        coid = norm(row[idx.get("cooperation_id", -1)]) if idx.get("cooperation_id") is not None else None
        chat = norm(row[idx.get("openChatId", -1)]) if idx.get("openChatId") is not None else None
        s = norm(row[idx.get("status", -1)]) if idx.get("status") is not None else None
        try:
            status = Status(s) if s else Status.UNKNOWN_ERROR
        except Exception:
            status = Status.UNKNOWN_ERROR
        ecode = norm(row[idx.get("error_code", -1)]) if idx.get("error_code") is not None else None
        emsg = norm(row[idx.get("error_message", -1)]) if idx.get("error_message") is not None else None

        if cn not in mapping:
            order.append(cn)
        mapping[cn] = ResultRow(
            contract_number=cn,
            contract_id=cid,
            cooperation_id=coid,
            openChatId=chat,
            status=status,
            error_code=ecode,
            error_message=emsg,
        )

    return order, mapping
